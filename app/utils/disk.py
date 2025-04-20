import openpyxl
import yadisk
import pandas as pd
import requests

from io import BytesIO

class YandexDisk:
    def __init__(self, id, secret, token):
        self.id = id
        self.secret = secret
        self.token = token
        self.client = None

    async def init_client(self):
        if self.client is None:
            self.client = yadisk.AsyncClient(id=self.id, secret=self.secret, token=self.token)

    async def read_file(self, file_path):
        await self.init_client()
        # Получаем ссылку на скачивание файла
        download_link = await self.client.get_download_link(file_path)

        # Получаем содержимое файла по ссылке
        response = requests.get(download_link)

        # Читаем Excel-файл из содержимого
        wb = openpyxl.load_workbook(BytesIO(response.content))

        return wb.active

    async def get_sources(self, file_path):
        sheet = await self.read_file(file_path)

        data = []
        for row in sheet.rows:
            row_data = [cell.value for cell in row if cell.value is not None]
            if len(row_data) >= 2 and row_data[1] is not None:
                data.append(row_data)

        df = pd.DataFrame(data)

        # Преобразуем DataFrame в список списков
        result = df.values.tolist() # название + ссылка

        return result

    
    async def get_ban_categories(self, file_path):
        sheet = await self.read_file(file_path)

        # Создаем пустой список для слов
        words = []

        # Итерируем по строкам листа
        for row in range(1, sheet.max_row + 1):
            # Получаем значение ячейки в текущей строке
            cell_value = sheet.cell(row=row, column=1).value

            # Если ячейка не пуста, добавляем значение в список слов
            if cell_value is not None:
                words.append(str(cell_value).strip())

        return words
    
    async def delete_source(self, file_path, row_number):
        try:
            # Получаем информацию о файле
            info = await self.client.get_meta(file_path)

            # Проверяем существование файла
            if info is None:
                raise FileNotFoundError(f"Файл {file_path} не существует на Яндекс.Диске")

            sheet = await self.read_file(file_path)

            # Проверяем значение sheet
            if sheet is None:
                raise ValueError("Не удалось прочитать файл")

            if row_number < 1 or row_number > sheet.max_row:
                raise ValueError("Номер строки находится вне допустимого диапазона")

            sheet.delete_rows(row_number)

            # Сохраняем обновленный лист в файле
            with BytesIO() as output:
                sheet.parent.save(output)
                output.seek(0)
                res = await self.client.upload(output, file_path, overwrite=True)

                if res:
                    return True
                
        except Exception as e:
            print(f"Ошибка: {e}")

    async def add_source(self, file_path, name, link):
        try:
            # Получаем информацию о файле
            info = await self.client.get_meta(file_path)

            # Проверяем существование файла
            if info is None:
                raise FileNotFoundError(f"Файл {file_path} не существует на Яндекс.Диске")

            # Скачиваем файл
            download_link = await self.client.get_download_link(file_path)
            response = requests.get(download_link)
            with BytesIO(response.content) as file:
                # Загружаем файл в openpyxl
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Добавляем новую строку
                sheet.append([name, link])

                # Сохраняем обновленный файл
                with BytesIO() as output:
                    wb.save(output)
                    output.seek(0)
                    res = await self.client.upload(output, file_path, overwrite=True)

                    return res
                
        except Exception as e:
            print(f"Ошибка: {e}")

    async def add_new_categories(self, file_path, lists):
        try:
            # Получаем информацию о файле
            info = await self.client.get_meta(file_path)

            # Проверяем существование файла
            if info is None:
                raise FileNotFoundError(f"Файл {file_path} не существует на Яндекс.Диске")

            # Скачиваем файл
            download_link = await self.client.get_download_link(file_path)
            response = requests.get(download_link)
            with BytesIO(response.content) as file:
                # Загружаем файл в openpyxl
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Добавляем новую строку
                for category in lists:
                    sheet.append([category])

                # Сохраняем обновленный файл
                with BytesIO() as output:
                    wb.save(output)
                    output.seek(0)
                    res = await self.client.upload(output, file_path, overwrite=True)

                    return res
                
        except Exception as e:
            print(f"Ошибка: {e}")